"""
Génération des exports PDF et XLSX de la liste d'action (US-E4).
"""
from __future__ import annotations
import io
from datetime import datetime
from typing import Any

# ── XLSX via openpyxl ────────────────────────────────────────────────────────

def generer_xlsx(lignes: list[dict[str, Any]], nom_officine: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COULEURS_STATUT = {
        "RUPTURE":  "FFCDD2",  # rouge clair
        "CRITIQUE": "FFE0B2",  # orange clair
        "COMMANDER":"FFF9C4",  # jaune clair
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Liste d'action"

    # Titre
    ws.merge_cells("A1:H1")
    ws["A1"] = f"SAD OFFICINE — Liste d'action — {nom_officine}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])  # ligne vide

    # En-têtes
    entetes = ["Code", "Désignation", "Classe", "Stock actuel", "Statut",
               "Qté à commander", "Valeur (FCFA)", "Décision"]
    ws.append(entetes)
    header_row = ws.max_row
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, _ in enumerate(entetes, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2D6A4F")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Données
    for ligne in lignes:
        row = [
            ligne["code"],
            ligne["designation"],
            ligne.get("classe") or "—",
            ligne["stock_actuel"],
            ligne["statut"],
            ligne["qte_a_commander"],
            round(ligne["valeur_fcfa"], 0),
            ligne["texte_decision"],
        ]
        ws.append(row)
        data_row = ws.max_row
        couleur = COULEURS_STATUT.get(ligne["statut"], "FFFFFF")
        for col in range(1, len(entetes) + 1):
            cell = ws.cell(row=data_row, column=col)
            cell.fill = PatternFill(fill_type="solid", fgColor=couleur)
            cell.border = border
            if col in (4, 6, 7):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"

    # Largeurs de colonnes
    largeurs = [14, 35, 8, 14, 12, 16, 16, 55]
    for i, larg in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF via reportlab ────────────────────────────────────────────────────────

def generer_pdf(lignes: list[dict[str, Any]], nom_officine: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )

    COULEURS_STATUT = {
        "RUPTURE":  colors.HexColor("#FFCDD2"),
        "CRITIQUE": colors.HexColor("#FFE0B2"),
        "COMMANDER":colors.HexColor("#FFF9C4"),
    }
    VERT = colors.HexColor("#2D6A4F")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle("titre", parent=styles["Title"], fontSize=14, textColor=VERT)
    sous_titre_style = ParagraphStyle("sous", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=9)

    elements = [
        Paragraph(f"SAD OFFICINE — Liste d'action", titre_style),
        Paragraph(f"{nom_officine} — {datetime.now().strftime('%d/%m/%Y à %H:%M')}", sous_titre_style),
        Spacer(1, 6*mm),
    ]

    # Table
    entetes = ["Code", "Désignation", "Cl.", "Stock", "Statut", "Qté", "Valeur FCFA", "Décision"]
    data = [entetes]
    for l in lignes:
        data.append([
            l["code"],
            Paragraph(l["designation"], cell_style),
            l.get("classe") or "—",
            str(int(l["stock_actuel"])),
            l["statut"],
            str(int(l["qte_a_commander"])),
            f"{int(l['valeur_fcfa']):,}".replace(",", " "),
            Paragraph(l["texte_decision"], cell_style),
        ])

    col_widths = [25*mm, 55*mm, 10*mm, 16*mm, 22*mm, 12*mm, 25*mm, 90*mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND",   (0, 0), (-1, 0), VERT),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 8),
        ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
        ("FONTSIZE",     (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID",         (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 3),
    ]

    # Couleur par statut sur les lignes de données
    for i, ligne in enumerate(lignes, start=1):
        couleur = COULEURS_STATUT.get(ligne["statut"])
        if couleur:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), couleur))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()
